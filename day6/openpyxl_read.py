from openpyxl import load_workbook

wb = load_workbook('./2024-winter-python/day6/3000words.xlsx',
                   read_only=False,  # 읽기 전용(읽기 전용에 최적화되어 파일을 불러온다)
                   # False면 셀안 공식을 가져오고 True면 공식 적용된 값만을 불러온다.
                   data_only=True, # True: 수식이 아닌 값을 가져옴
                   )
sheet1 = wb['Sheet1']  # Sheet1(이름) 시트를 가져오기.

print(sheet1['B3'].value)  # B3 데이터 출력
print(sheet1.cell(1,3).value) # (3,1) 좌표 데이터 출력

get_cells = sheet1['A2' : 'D2']  # A2부터 D2까지의 데이터를 가져오기
for row in get_cells:
    for cell in row:
        print(cell.value)  # 이중반복문을 통한 출력.

# 모든 행과 열 출력 + list 이용
all_values = []
for row in sheet1.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)